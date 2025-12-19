#!/usr/bin/env python3
"""
Export reconciliation data to XLSX file.
Reads order IDs from stdin (JSON format) and outputs XLSX to stdout.
"""
import os
import sys
import json
from io import BytesIO
import psycopg2
import openpyxl
from openpyxl import Workbook

def get_db_connection():
    """Create database connection using DATABASE_URL environment variable."""
    database_url = os.environ.get('DATABASE_URL')
    if not database_url:
        raise Exception("DATABASE_URL environment variable not set")
    return psycopg2.connect(database_url)

def export_orders(order_ids):
    """Generate XLS file bytes for the given order IDs."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    placeholders = ','.join(['%s'] * len(order_ids))
    query = f"""
        SELECT order_id, amount_total_fee, order_bank_reference
        FROM orders
        WHERE order_id IN ({placeholders})
        ORDER BY order_id
    """
    cursor.execute(query, order_ids)
    rows = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    
    ws['A1'] = 'Order Number'
    ws['B1'] = 'Order Value'
    ws['C1'] = 'Bank Reference'
    
    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].font = openpyxl.styles.Font(bold=True)
    
    for row_idx, row in enumerate(rows, start=2):
        ws[f'A{row_idx}'] = row[0]
        ws[f'B{row_idx}'] = float(row[1]) if row[1] else 0
        ws[f'C{row_idx}'] = row[2] or ''
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def main():
    try:
        input_data = json.loads(sys.stdin.read())
        order_ids = input_data.get('orderIds', [])
        
        if not order_ids:
            print(json.dumps({"error": "No order IDs provided"}), file=sys.stderr)
            sys.exit(1)
        
        xls_bytes = export_orders(order_ids)
        sys.stdout.buffer.write(xls_bytes)
        
    except Exception as e:
        print(json.dumps({"error": str(e)}), file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
